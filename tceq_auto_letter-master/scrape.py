import openpyxl
import pickle

print("Loading Excel file")
# file_1Y = '1Y_2018_SPP.xlsx'
file_3Y = '3Y_2017_SPP.xlsx'
# wb_1Y = openpyxl.load_workbook(file_1Y)
wb_3Y = openpyxl.load_workbook(file_3Y)
# sheet_1Y = wb_1Y['COPY']
sheet_3Y = wb_3Y['SPP_MASTER']
print("Excel file loaded")


def list_pws(sheet):
    # PULLS DATA FROM EXCEL FILE AND CONVERTS TO A PYTHON DICTIONARY OBJECT
    print("SCRAPING EXCEL FILE: {} FOR DATA".format(file_3Y))
    dct = {}
    samples = []
    for i in range(2, 39150):
        pws_id = sheet.cell(row=i, column=1).value
        name = sheet.cell(row=i, column=2).value
        pop = sheet.cell(row=i, column=7).value
        sample_id = sheet.cell(row=i, column=8).value
        loc = sheet.cell(row=i, column=9).value
        activity = sheet.cell(row=i, column=11).value
        tier = sheet.cell(row=i, column=15).value
        update_by = sheet.cell(row=i, column=19).value
        creation_date = sheet.cell(row=i, column=22).value
        if creation_date is not None:
            creation_date = creation_date.strftime('%x')
        sample = sample_id, activity, loc, tier, update_by, creation_date
        samples.append(sample)
        new_row = i + 1
        if pws_id != sheet.cell(row=new_row, column=1).value:
            dct[pws_id] = [pws_id, name, pop, samples]
            samples = []
    print("CONVERTING DATA TO DICTIONARY OBJECT")
    return dct


""" Function list_pws returns a dictionary object:
    dct = {pws_id: [
        pws_id,
        name,
        pop,
        [[sample_id,
          activity,
          loc,tier,
          update_by,
          creation_date],
         ...]],
        ...}"""

# CONVERTS PYTHON DICTIONARY OBJECT TO AN EXTERNAL PICKLE FILE
pws = list_pws(sheet_3Y)
print("PICKLING DICTIONARY OBJECT")
f = open("pws_list.pkl", "wb")
pickle.dump(pws, f)
f.close()
print("PICKLING COMPLETE")
